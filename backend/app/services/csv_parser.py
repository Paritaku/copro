import csv
import re
from typing import List, Optional
from app.models.models import Lot, Floor, ImportedData
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Side, Border
from openpyxl.utils import column_index_from_string, get_column_letter
from io import BytesIO
from decimal import ROUND_HALF_UP, Decimal


class CSVParser:
    """Parser pour fichiers CSV de titres fonciers"""
    unicode = "\u1D43"
    
    def __init__(self, delimiter=";"):
        self.delimiter = delimiter
    
    def parse_file(self, file_path: str) -> ImportedData:
        """Parse un fichier CSV de titre foncier"""
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=self.delimiter)
            rows = [row for row in reader]
        
        return self._parse_rows(rows)
    
    def parse_content(self, content: str) -> ImportedData:
        """Parse le contenu CSV en string"""
        rows = []
        for line in content.split('\n'):
            rows.append(line.split(self.delimiter))
        
        return self._parse_rows(rows)
    
    def _parse_rows(self, rows: List[List[str]]) -> ImportedData:
        """Parse les lignes du CSV"""
        titre_foncier = ""
        etages: List[Floor] = []
        
        i = 0

        # Extraire le titre foncier
        while i < len(rows):
            row = rows[i]
            if len(row) > 1 and "Titre foncier" in row[1]:
                # Ex: "Modification successives du Titre foncier  :154311 /05"
                titre_foncier = row[1].split(":")[-1].strip()
                i += 1
                break
            i += 1
        
        # Sauter jusqu'aux données réelles
        while i < len(rows):
            row = rows[i]
            if len(row) > 1 and row[1] and "Propriété dite" in row[1]:
                i += 3
                break
            i += 1

        # Parser les étages et lots
        while i < len(rows):
            row = rows[i]
            if len(row) > 0 and row[1] and ":" in row[1] and any(c.isalpha() for c in row[1]):
                parts = row[1].split(":")
                etage_name = parts[0].strip()  # "Rez-de-chaussée"
                cotes = parts[1].strip() if len(parts) > 1 else ""  # "Des côtes +0.10m et +1,10m à la côte 4,10m"
                    
                lots: List[Lot] = []
                i += 1
               
                # Parser les lots de cet étage
                total_surf_int = 0
                total_surf_surp = 0
                
                while i < len(rows):
                    row = rows[i]

                    # Arrêt si nouvel étage
                    if len(row) > 0 and row[1] and ":" in row[1] and any(c.isalpha() for c in row[1]):
                        break
                    
                    # Arrêt si fin des données
                    if not row or not any(cell.strip() for cell in row):
                        i += 1
                        continue
                    
                    # Total row
                    if len(row) > 1 and "Total" in row[2]:
                        i += 1
                        break
                    
                    # Parser un lot
                    if len(row) > 3 and row[5]: # Propriété et Surface interieure du titre
                        lot = self._parse_lot(row)
                        if lot:
                            lots.append(lot)
                            total_surf_int += lot.surface_interieure or 0
                            total_surf_surp += lot.surface_avec_surplomb or 0
                    
                    i += 1

                if lots:
                    floor = Floor(
                        nom=etage_name,
                        cotes=cotes,
                        lots=lots,
                        total_surface_interieure=total_surf_int,
                        total_surface_avec_surplomb=total_surf_surp
                    )
                    etages.append(floor)
            else:
                i += 1

        return ImportedData(
            titre_foncier=titre_foncier,
            etages=etages
        )
    
    def _parse_lot(self, row: List[str]) -> Optional[Lot]:
        """Parse une ligne représentant un lot"""

        try:
            propriete = row[1].strip() if len(row) > 0 else ""
            titre_num = row[2].strip() if len(row) > 1 else ""

            indice_privative = row[3].strip() if len(row) > 3 else None
            indice_commune = row[4].strip() if len(row) > 4 else None

            surface_interieure = self._parse_float(row[5]) if len(row) > 5 else 0
            surface_avec_surplomb = self._parse_float(row[6]) if len(row) > 6 else 0
            
            consistance = row[7].strip() if len(row) > 7 else ""
            observations = row[8].strip() if len(row) > 8 else None
            
            # Valider que c'est un lot valide (au moins propriete et indice)
            if indice_privative or indice_commune:
                return Lot(
                    propriete=propriete,
                    titre_num=titre_num,
                    indice_privative=indice_privative,
                    indice_commune=indice_commune,
                    surface_interieure=surface_interieure,
                    surface_avec_surplomb=surface_avec_surplomb,
                    consistance=consistance,
                    observations=observations
                )
        except Exception as e:
            print(f"Erreur parsing lot: {e}")
        
        return None
    
    def _parse_float(self, value: str) -> Optional[float]:
        """Parse une valeur float en gérant les formats différents"""
        if not value or not value.strip():
            return None
        value = value.strip()
        try:
            return float(value)
        except ValueError:
            return None

    def _style_cell(self, cell, font: Font =  None, alignment: Alignment = None, border: Border = None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    def _create_font(
        self,
        name: str, 
        size: int = 12, 
        bold: bool = False,
        underline = None
    ) -> Font:
        return Font(
            name = name,
            size = size,
            bold= bold,
            underline= underline
        )
        
    def _create_arial_font(
        self,
        size: int = 12,
        bold: bool = False
    ) -> Font:
        return self._create_font("Arial", size, bold)
    
    def _create_arial_narrow_font(
        self,
        size: int = 12,
        bold: bool = False
    ) -> Font:
        return self._create_font("Arial Narrow", size, bold)
    
    def _create_times_new_roman_font(
        self,
        size: int = 12,
        bold: bool = False,
        underline = None
    ) -> Font:
        return self._create_font("Times New Roman", size, bold, underline)
    
    def _fully_centered(self, wrap_text = False) -> Alignment:
        return Alignment(vertical="center", horizontal="center", wrap_text=wrap_text)
    
    def _solid_black_border(self, style: str = "thin"): 
        borderSide = Side(style = style, border_style=None, color='FF000000')
        return Border(left=borderSide, top=borderSide, bottom=borderSide, right=borderSide)
        
    def generer_xlxs_voix(self, data: ImportedData):
        wb = Workbook()
        ws = wb.active
        ws.title = "Voix"
        
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 30
        
        fontArial =  Font(
            name="Arial",
            size=12,
        )
        fontArialBold = Font(
            name="Arial",
            size=12,
            bold=True
        )

        texte = "Règlement de coproprieté"
        titre_foncier = f"TF {data.titre_foncier}"
        propriete_dite = f"Proprieté dite : {data.etages[0].lots[0].propriete.split("-")[0]}"
        
        ws.merge_cells("B2:H2")
        ws.merge_cells("B5:H5")
        ws.merge_cells("B7:H13")
        ws.merge_cells("B14:H14")
        ws.merge_cells("B16:H18")
        ws.merge_cells("B19:H19")
        
        borderSide = Side(style="thick",border_style=None, color='FF000000')
        border = Border(left=borderSide, top=borderSide, bottom=borderSide, right=borderSide)
        
        for col in range(2, 9): # B=2 → H=8
            for row in [14,19]:
                ws.cell(row=row, column=col).border = border
        
        cell = ws["B2"]
        cell.value = f"{texte}      {titre_foncier}     {propriete_dite}"
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(
            underline="single", 
            size=12, 
            name="Times New Roman")
        
        cell = ws["B5"]
        cell.value = "2-LE NOMBRE DE VOIX DES COPROPRIETAIRES"
        cell.font = fontArialBold
        cell.alignment = Alignment(vertical="center")
        
        cell = ws["B7"]
        cell.value = (
                "Le nombre des voix des copropriétaires est calculé en fonction de la surface de la partie privative qui leur est attribuée\n"
                "Si = surface de la partie privative i.\n"
                "NVi =Nombre de voix du propriétaire de la partie privative i.\n"
                "S = Total des surfaces des parties privatives de l’immeuble.\n"
                "N = le nombre de millième (1.000) ou de dix-millième (10.000) " 
        )
        cell.font = fontArial
        cell.alignment= Alignment(vertical="top", horizontal="left",wrap_text=True)
        
        cell = ws["B14"]
        cell.value = "Nvi = (Si/S) x 100"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = fontArialBold
        
        cell = ws["B16"]
        cell.value = (
            "En d’autres termes, le nombre de voix peut  être déduit à partir du tantième d’indivision. " 
            "En effet, soit T.Ii le tantième d’indivision du propriétaire de la partie privative i. le nombre de voix ce dernier est :"
        )
        cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
        cell.font = fontArial
        
        cell = ws["B19"]
        cell.value = "Nvi = (T.Ii/N) x 100"
        cell.font = fontArialBold
        cell.alignment = Alignment(horizontal="center")
        
        
        for col in range(2,8):
            ws.merge_cells(start_row=21, end_row=23, start_column=col, end_column=col)
            
        for c in range (column_index_from_string("B"), column_index_from_string("G")+1):
            cell = ws.cell(column=c, row=21)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.font = fontArialBold
            
        for c in range (column_index_from_string("B"), column_index_from_string("G")+1):
            for r in range(21,24):
                cell = ws.cell(column=c, row=r)
                cell.border = border
            
        cell = ws["B21"]
        cell.value = "N° d'ordre"
        
        cell = ws["C21"]
        cell.value = "Indices de la Partie Privative"
        
        cell = ws["D21"]
        cell.value = "Niveau"
        
        cell = ws["E21"]
        cell.value = "Consistance"
        
        cell = ws["F21"]
        cell.value = "Si  (m²)"
        
        cell = ws["G21"]
        cell.value = "NVi  (%)"
        
        for i in range (21,24):
            ws.row_dimensions[i].height = 25

        surface_totale_partie_privative = 0
        somme_des_nvi = 0
        num_ordre = 1
        current_line = 24
        
        for etage in data.etages:
             for lot in etage.lots:
                if(lot.indice_privative):
                    surface_totale_partie_privative += lot.surface_avec_surplomb
        
        for etage in data.etages:
            start_merge = current_line
            nb_de_lignes_a_merger = 0
            
            for lot in etage.lots:
                if lot.indice_privative: 
                    ws.row_dimensions[current_line].height = 30
                    cell = ws[f"B{current_line}"]
                    cell.value = num_ordre
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    cell.font = fontArial
                    
                    cell = ws[f"C{current_line}"] 
                    cell.value = lot.indice_privative
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    cell.font = fontArial
                    
                    cell = ws[f"E{current_line}"] 
                    cell.value = lot.consistance
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    cell.font = fontArial
                    
                    cell = ws[f"F{current_line}"] 
                    cell.value = lot.surface_avec_surplomb
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    cell.font = fontArialBold
                    
                    cell = ws[f"G{current_line}"]
                    nvi = Decimal(lot.surface_avec_surplomb) / Decimal (surface_totale_partie_privative) * Decimal(100)
                    cell.value = nvi.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    cell.font = fontArial
                    
                    somme_des_nvi += nvi
                    nb_de_lignes_a_merger += 1
                    current_line += 1
                    num_ordre += 1
                    
    
            if any(lot.indice_privative for lot in etage.lots):
                cell = ws[f"D{start_merge}"] 
                cell.value = etage.nom
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.font = fontArial
                ws.merge_cells(f"D{start_merge}:D{start_merge + nb_de_lignes_a_merger-1}")
              
        
        ws.merge_cells(f"B{current_line}:E{current_line}")
        current_cell = ws[f"B{current_line}"]
        current_cell.value = "Total"
        current_cell.alignment = Alignment(vertical= "center", horizontal="center")
        current_cell.font = fontArial
        ws.row_dimensions[current_line].height = 30
        
        current_cell = ws[f"F{current_line}"]
        current_cell.value = surface_totale_partie_privative
        current_cell.font = fontArialBold
        current_cell.alignment = Alignment(vertical= "center", horizontal="center")
        
        current_cell = ws[f"G{current_line}"]
        current_cell.value = somme_des_nvi
        current_cell.font = fontArialBold
        current_cell.alignment = Alignment(vertical= "center", horizontal="center")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def generer_xlxs_quotation(self, data: ImportedData):
        wb = Workbook()
        ws = wb.active
        ws.title = "Quot P CH2"
      
        # ========================== FONT ==========================
        fontTimes16Bold = self._create_times_new_roman_font(16, True)
        fontTimes14Bold = self._create_times_new_roman_font(14, True)
        fontTimes14 = self._create_times_new_roman_font(14, False)
        fontArial12Bold = self._create_arial_font(12, True)
        fontArial12 = self._create_arial_font(12, False)
        fontArial14RedBold = Font(name="Arial", size=14, bold=True, color="FF0000")
        fontArial14Bold = self._create_arial_font(14, True)

        # ========================== FONT ========================== 
        
        # ========================== Headers ==========================
        for i in range(2,4):
            ws.row_dimensions[i].height = 30
            
        ws.merge_cells("B4:H4")
        cell = ws["B4"]
        cell.value = "II-QUOTS-PARTS ET VOIX"
        cell.font = fontTimes16Bold        
        for i in range(2, 11):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 133 / 7
            if(i == 10):
                ws.column_dimensions[letter].width = 50
        cell.alignment = Alignment(vertical = "center")
        
        
        ws.merge_cells("B5:H5")
        cell = ws["B5"]
        cell.value = "1- TABLEAU DE REPARTITION DES QUOTS-PARTS ET DIMILIEME D'INDIVISION"
        cell.font = fontTimes16Bold
        cell.alignment = Alignment(vertical = "center")

        ws.merge_cells("B7:C8")
        cell = ws["B7"]
        cell.value = "Indices des parties"
        cell.font = fontTimes16Bold
        cell.alignment = self._fully_centered()
        
        for row in ws["B7:C8"]:
            for cell in row:
                cell.border = self._solid_black_border("thin")
                
        cell = ws["B9"]
        cell.value = "Privative"
        cell.font = fontTimes14
        cell.alignment = self._fully_centered()
        cell.border = self._solid_black_border("thin")
        
        cell = ws["C9"]
        cell.value = "Commune"
        cell.font = fontTimes14
        cell.alignment = self._fully_centered()
        cell.border = self._solid_black_border("thin")
        
        ws.merge_cells("D7:E9")
        cell = ws["D7"]
        cell.value = "Consistance"
        cell.font = fontTimes16Bold
        cell.alignment = self._fully_centered()
        
        for row in ws["D7:E9"]:
            for cell in row:
                cell.border = self._solid_black_border("thin")
        
        ws.merge_cells("F7:G8")
        cell = ws["F7"]
        cell.value = "Surface en m²"
        cell.font = fontTimes16Bold
        cell.alignment = self._fully_centered()
        cell.border = self._solid_black_border("thin")
        
        for row in ws["F7:G8"]:
            for cell in row:
                cell.border = self._solid_black_border("thin")
        
        cell = ws["F9"]
        cell.font = fontTimes14
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.border = self._solid_black_border("thin")
        cell.value = "Intérieure du Titre"
        
        cell = ws["G9"]
        cell.font = fontTimes14
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.border = self._solid_black_border("thin")
        cell.value = "Total avec surplomb"
        
        ws.merge_cells("H7:H9")
        cell = ws["H7"]
        cell.font = fontTimes14
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.border = self._solid_black_border("thin")
        cell.value = "Qute-Part du Terrain(m²)"
        
        for row in ws["H7:H9"]:
            for cell in row:
                cell.border = self._solid_black_border("thin")
        
        
        ws.merge_cells("I7:I9")
        cell = ws["I7"]
        cell.font = fontTimes14
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.border = self._solid_black_border("thin")
        cell.value = "Part D'indivision En (1/10000)"
        
        for row in ws["I7:I9"]:
            for cell in row:
                cell.border = self._solid_black_border("thin")
        
        
        ws.merge_cells("J7:J9")
        cell = ws["J7"]
        cell.font = fontTimes16Bold
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.border = self._solid_black_border("thin")
        cell.value = "Observations"
        
        for row in ws["J7:J9"]:
            for cell in row:
                cell.border = self._solid_black_border("thin")
                
        current_line = 10
        totaux_generaux_surface_interieure = 0
        totaux_generaux_surface_avec_surplomb = 0
        totaux_generaux_quots_parts = 0 
        totaux_generaux_part_d_indivision = 0
        
        for etage in data.etages:
            start_merge = current_line        
            ws.merge_cells(f"B{start_merge}:J{start_merge}")
            cell = ws[f"B{start_merge}"]
            cell.value = f"{etage.nom} : {etage.cotes}"
            cell.font = fontTimes16Bold
            cell.alignment = self._fully_centered()
            
            ws.row_dimensions[current_line].height = 30
            self._apply_border_to_range(ws=ws, type="thin", range=f"B{start_merge}:J{start_merge}")
            
            
            for lot in etage.lots:
                current_line += 1
                ws.row_dimensions[current_line].height = 30
                #Privative
                cell = ws[f"B{current_line}"]
                cell.font = fontArial12Bold
                cell.alignment = self._fully_centered()
                cell.value = lot.indice_privative.replace("a", self.unicode) if lot.indice_privative else ""
                
                 #Commune
                cell = ws[f"C{current_line}"]
                cell.font = fontArial12
                cell.alignment = self._fully_centered()
                cell.value = lot.indice_commune.replace("a", self.unicode) if lot.indice_commune else ""
                
                #Consistance
                ws.merge_cells(f"D{current_line}:E{current_line}")
                cell = ws[f"D{current_line}"]
                cell.value = lot.consistance
                cell.font =  fontArial12Bold if lot.indice_privative else fontTimes14
                cell.alignment = self._fully_centered()
                
               #Interieure du titre
                cell = ws[f"F{current_line}"]
                cell.font = fontArial12Bold
                cell.alignment = self._fully_centered()
                cell.value = lot.surface_interieure
                
                #Total avec surplomb du titre
                cell = ws[f"G{current_line}"]
                cell.font = fontArial12Bold
                cell.alignment = self._fully_centered()
                cell.value = lot.surface_avec_surplomb
                
                #Observations 
                cell = ws[f"J{current_line}"]
                cell.font = self._create_arial_narrow_font(10, False)
                cell.alignment = self._fully_centered()
                cell.value = lot.observations.replace("a", self.unicode) if lot.observations else ""
                
                if lot.indice_privative:
                    totaux_generaux_surface_interieure += lot.surface_interieure
                    totaux_generaux_surface_avec_surplomb += lot.surface_avec_surplomb
            
            #Total
            current_line += 1
            ws.row_dimensions[current_line].height = 30
            ws.merge_cells(f"B{current_line}:E{current_line}")
            cell = ws[f"B{current_line}"]
            cell.font = fontArial12Bold
            cell.alignment = self._fully_centered()
            cell.value = "Total"
            
            cell = ws[f"F{current_line}"]
            cell.font = fontArial12Bold
            cell.alignment = self._fully_centered()
            cell.value = etage.total_surface_interieure
            
            cell = ws[f"G{current_line}"]
            cell.font = fontArial12Bold
            cell.alignment = self._fully_centered()
            cell.value = etage.total_surface_avec_surplomb
            
            current_line += 1
            
        current_line = 10
        
        #Calcul quots-parts et part d'indivision
        for etage in data.etages: 
            current_line += 1
            totaux_quots_etage = Decimal(0)
            totaux_indivision_etage = Decimal(0)
            
            for lot in etage.lots:
                if(lot.indice_privative):
                    quot = Decimal(lot.surface_avec_surplomb) * Decimal(etage.total_surface_interieure) / Decimal(totaux_generaux_surface_avec_surplomb) 
                    indivision = Decimal(lot.surface_avec_surplomb)  * Decimal(10000) / Decimal(totaux_generaux_surface_avec_surplomb)
                
                    print(current_line, quot, indivision)
                
                    totaux_quots_etage += quot
                    totaux_indivision_etage += indivision
                
                    totaux_generaux_part_d_indivision += indivision
                    totaux_generaux_quots_parts += quot
                    
                    cell = ws[f"H{current_line}"]
                    cell.font = fontArial14Bold
                    cell.alignment = self._fully_centered()
                    cell.value = quot.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) 
                    
                    cell = ws[f"I{current_line}"]
                    cell.font = fontArial14Bold
                    cell.alignment = self._fully_centered()
                    cell.value = indivision.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                
                current_line += 1
                
            #Total etage
            cell = ws[f"H{current_line}"]
            cell.font = fontArial14RedBold
            cell.alignment = self._fully_centered()
            cell.value = totaux_quots_etage.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if totaux_quots_etage > 0 else ""
                    
            cell = ws[f"I{current_line}"]
            cell.font = fontArial14RedBold
            cell.alignment = self._fully_centered()
            cell.value = totaux_indivision_etage.quantize(Decimal("1"), rounding=ROUND_HALF_UP) if totaux_indivision_etage > 0 else ""
            
            current_line += 1
            
        # Totaux généraux
        ws.merge_cells(f"B{current_line}:E{current_line}")
        ws.row_dimensions[current_line].height = 30
        cell = ws[f"B{current_line}"]
        cell.font = fontArial12Bold
        cell.alignment = self._fully_centered()
        cell.value = "Totaux Généraux"
        self._apply_border_to_range(ws=ws, type="thin", range=f"B{current_line}:E{current_line}")
        cell.border = self._solid_black_border("thin")
                
        cell = ws[f"F{current_line}"]  
        cell.font = fontArial12Bold
        cell.alignment = self._fully_centered()
        cell.value = totaux_generaux_surface_interieure
        cell.border = self._solid_black_border("thin")
        
        cell = ws[f"G{current_line}"]  
        cell.font = fontArial12Bold
        cell.alignment = self._fully_centered()
        cell.value = totaux_generaux_surface_avec_surplomb
        cell.border = self._solid_black_border("thin")
        
        cell = ws[f"H{current_line}"]  
        cell.font = fontArial14Bold
        cell.alignment = self._fully_centered()
        cell.value = totaux_generaux_quots_parts
        cell.border = self._solid_black_border("thin")
        
        cell = ws[f"I{current_line}"]  
        cell.font = fontArial14Bold
        cell.alignment = self._fully_centered()
        cell.value = totaux_generaux_part_d_indivision.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        cell.border = self._solid_black_border("thin")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
        # ========================== Headers ==========================
        
    def _apply_border_to_range(self, ws, type: str = "thin", range: str= ""): 
        for row in ws[range]:
            for cell in row:
                cell.border = self._solid_black_border(type)
                
    def generer_xlxs_ta(self, data: ImportedData):
        wb = Workbook()
        ws = wb.active
        ws.title = "TA"
        
        for i in range (1,8):
            ws.column_dimensions[get_column_letter(i)].width = 25
            
            if i in range (6,8):
                ws.column_dimensions[get_column_letter(i)].width = 30
                
            if i in range (3,5):
                ws.column_dimensions[get_column_letter(i)].width = 15
        
        # =========FONT============
        arial14bold = self._create_arial_font(14, True)
        arial12bold = self._create_arial_font(12, True)
        arial12 = self._create_arial_font(12, False)
        arialNarrow12 = self._create_arial_narrow_font(12, False)
        
        # ==============TABLE TITLE ============
        ws.merge_cells("A3:H3")
        cell = ws["A3"]
        cell.font = arial14bold
        cell.value = f"Titre foncier : {data.titre_foncier}"
        
        ws.merge_cells("B4:G4")
        cell = ws["B4"]
        cell.font = arial14bold
        cell.value = "TABLEAU « A » DES CONTENANCES DE LA COPROPRIETE"
        
        
        # ========= HEADERS ================
        for i in range (1,8):
            letter = get_column_letter(i)  
            ws.merge_cells(f"{letter}6:{letter}8")
       
        # Propriete dite
        cell = ws["A6"]
        cell.value = "Propriété dite"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="A6:A8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
        # Indices des 
        cell = ws["B6"]
        cell.value = "Titre N°"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="B6:B8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
        # Indices des 
        cell = ws["C6"]
        cell.value = "Indices des P,P"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="C6:C8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
        # Surface en m² 
        cell = ws["D6"]
        cell.value = "Surface en m²"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="D6:D8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
        # Surface en m² 
        cell = ws["E6"]
        cell.value = "Situation"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="E6:E8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
        # Surface en m² 
        cell = ws["F6"]
        cell.value = "Consistance"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="F6:F8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
          # Surface en m² 
        cell = ws["G6"]
        cell.value = "Observations"
        cell.font = arial12bold
        self._apply_border_to_range(ws = ws, type="thin", range="G6:G8")
        cell.alignment = self._fully_centered(wrap_text=True)
        
        current_line = 8
        for etage in data.etages:
            merge_start = current_line +1
            for lot in etage.lots:
                if (lot.indice_privative):
                    current_line += 1
                    ws.row_dimensions[current_line].height = 20
                    cell = ws[f"A{current_line}"]
                    cell.value = lot.propriete
                    cell.alignment = self._fully_centered()
                    cell.font = arial12bold
                    
                    cell = ws[f"C{current_line}"]
                    cell.value = lot.indice_privative.replace("a", self.unicode)
                    cell.alignment = self._fully_centered()
                    cell.font = arial12
                    
                    cell = ws[f"D{current_line}"]
                    cell.value = lot.surface_avec_surplomb
                    cell.alignment = self._fully_centered()
                    cell.font = arial12bold
                    
                    cell = ws[f"F{current_line}"]
                    cell.value = lot.consistance
                    cell.alignment = self._fully_centered()
                    cell.font = arial12
                    
                    cell = ws[f"G{current_line}"]
                    cell.value = lot.observations.replace("m2","m²").replace("a", self.unicode)
                    cell.alignment = self._fully_centered()
                    cell.font = arialNarrow12
                    
            if any(lot.indice_privative for lot in etage.lots):
                ws.merge_cells(f"E{merge_start}:E{current_line}")
                cell = ws[f"E{merge_start}"]
                cell.value = etage.nom
                cell.alignment = self._fully_centered()
                cell.font = arial12
                
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    def generer_excel_tr_n(self, data: ImportedData):
        wb = Workbook()
        ws = wb.active
        ws.title = "TR-N"
        
        for i in range (2,11):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 20
            
        # =============== FONT ===============
        arial12bold = self._create_arial_font(12, True)
        times12BoldUnderlined = self._create_times_new_roman_font(14, True, "single")
        arial14 = self._create_arial_font(14, False)
        arial12= self._create_arial_font(12, False)
        
        # Headers
        ws.merge_cells("B2:D6")
        cell = ws["B2"]
        cell.font = arial12bold
        self._apply_border_to_range(ws=ws, type="thin", range="B2:D6")
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.value = "ROYAUME DU MAROC\nAgence Nationale de la conservation Foncière Du Cadastre et de la Cartographie\n Conservation Foncière Meknés"

        ws.merge_cells("E2:F6")
        cell = ws["E2"]
        cell.font = arial12bold
        self._apply_border_to_range(ws=ws, type="thin", range="E2:F6")
        cell.alignment = self._fully_centered()
        cell.value = "N° de la pièce : 1/2"
        
        for i in range(2,7):
            ws.merge_cells(f"G{i}:J{i}")
            ws.row_dimensions[i].height = 25
            letter = get_column_letter(i)
            ws[f"G{i}"].font = arial12bold
            ws[f"G{i}"].alignment = Alignment(vertical="center")
            self._apply_border_to_range(ws=ws, type="thin", range=f"G{i}:J{i}")
            
        cell = ws["G2"]
        cell.value = f"Titre Foncier : {data.titre_foncier}"

        cell = ws["G3"]
        cell.value = "IGT : Ahmed El Hmidi"
 
        cell = ws["G4"]
        cell.value = "Carnet / Bon : "
        
        cell = ws["G5"]
        cell.value = "Date de délivrance : "
        
        cell = ws["G6"]
        cell.value = "Zoning : "
        
        for i in range (8, 10):
            letter = get_column_letter(i)
            ws.merge_cells(f"D{i}:J{i}")
            cell = ws[f"D{i}"]
            cell.font = times12BoldUnderlined
            cell.alignment = self._fully_centered()
            ws.row_dimensions[i].height = 30
        
        cell = ws["D8"]
        cell.value = "Copropriété" 
       
        cell = ws["D9"]
        cell.value = "Tableau détaillé des superficies par niveau"
        
        
        ws.merge_cells("B10:B11")
        self._apply_border_to_range(ws=ws, type="thin", range="B10:B11")
        ws.merge_cells("C10:E11")
        self._apply_border_to_range(ws=ws, type="thin", range="C10:E11")
        ws.merge_cells("F10:I10")
        self._apply_border_to_range(ws=ws, type="thin", range="F10:I10")
        ws.merge_cells("J10:J11")
        self._apply_border_to_range(ws=ws, type="thin", range="J10:J11")
        

        cell = ws["B10"]
        cell.value = "Niveau"
        cell.alignment = self._fully_centered()
        cell.font = arial14
        
        cell = ws["C10"]
        cell.value = "Consistance"
        cell.alignment = self._fully_centered()
        cell.font = arial14
        
        cell = ws["F10"]
        cell.value = "Superficies (m²)"
        cell.alignment = self._fully_centered()
        cell.font = arial14
        
        cell = ws["J10"]
        cell.value = "Total"
        cell.alignment = self._fully_centered()
        cell.font = arial14
        
        cell = ws["F11"]
        cell.value = "Superficie des P,P(1)"
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.font = arial14
        cell.border = self._solid_black_border(style="thin")
        
        cell = ws["G11"]
        cell.value = "Cours"
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.font = arial14
        cell.border = self._solid_black_border(style="thin")
        
        cell = ws["H11"]
        cell.value = "Balcon"
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.border = self._solid_black_border(style="thin")
        cell.font = arial14
        
        cell = ws["I11"]
        cell.value = "Terrasse"
        cell.alignment = self._fully_centered(wrap_text=True)
        cell.font = arial14
        cell.border = self._solid_black_border(style="thin")
        
        current_line = 11
        for etage in data.etages:
            start_merge = current_line + 1
            nb_ligne_merge = 0
            print(f"==============={etage.nom}=====================s")
            commerce_lots = [lot for lot in etage.lots if "commerc" in lot.consistance.lower() and lot.indice_privative]
            appartements_lots = [lot for lot in etage.lots if "appartement" in lot.consistance.lower() and lot.indice_privative]

            if(commerce_lots.__len__() > 0):
                current_line += 1
                print(current_line)
                nb_ligne_merge += 1
                ws.row_dimensions[current_line].height = 30
                ws.merge_cells(f"C{current_line}:E{current_line}")
                cell = ws[f"C{current_line}"]
                cell.value = "(Commerces)"
                cell.alignment = self._fully_centered()
                cell.font = arial12
                
                cell = ws[f"J{current_line}"]
                cell.value = f"{sum(lot.surface_avec_surplomb for lot in commerce_lots)} m²"
                cell.alignment = self._fully_centered()
                cell.font = arial12
                
            if(appartements_lots.__len__() > 0):
                current_line += 1
                nb_ligne_merge += 1
                ws.row_dimensions[current_line].height = 30
                ws.merge_cells(f"C{current_line}:E{current_line}")
                cell = ws[f"C{current_line}"]
                cell.value = "Appartements(habitation)"
                cell.alignment = self._fully_centered()
                cell.font = arial12
                print(current_line)
                
                cell = ws[f"J{current_line}"]
                cell.value = f"{sum(lot.surface_avec_surplomb for lot in commerce_lots)} m²"
                cell.alignment = self._fully_centered()
                cell.font = arial12
                            
            if(nb_ligne_merge >= 2):
                ws.merge_cells(f"B{start_merge}:B{start_merge+nb_ligne_merge-1}")
                
            cell = ws[f"B{start_merge}"]
            cell.value = etage.nom
            cell.alignment = self._fully_centered()
            cell.font = arial12bold  
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer